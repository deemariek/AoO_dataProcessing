import os
import datetime
from openpyxl import load_workbook, cell


months = {1: 'Jan',  2: 'Feb',   3: 'Mar',   4: 'Apr',
          5: 'May',  6: 'June',  7: 'July',  8: 'Aug',
          9: 'Sep', 10: 'Oct',  11: 'Nov',  12: 'Dec'}

# every month update these values to reflect the file names
studyMonth = months[2]
roundNumber = '10'

countryCodes = ['IRQ', 'JOR', 'LBN', 'TUR']
filePath = '..\\AoO_Round{0}_Data\\dataCleaning'.format(roundNumber)  
baseFileName = '{0}_KI_Village_Level_Monitoring_Tool_{1}16.xlsx'  
outputFile = '{0}_KI_Village_Level_Monitoring_Tool_{1}16_coded.csv'
## maybe pull the xlsx files from this folder directly as opposed to assuming their naming conventions.

# #########################################################################

def addCountryCodeColumn(studyMonth, countryCodes):
    print "Appending country code to files"
    for country in countryCodes:
        try:         
            fileName = baseFileName.format(country, studyMonth)
            fullFilePath = os.path.join(filePath, fileName)
            if os.path.isfile(fullFilePath):
                workbook = load_workbook(fullFilePath)
                first_sheet = workbook.get_sheet_names()[0]
                worksheet = workbook.get_sheet_by_name(first_sheet)
                highestRow = worksheet.get_highest_row()
                highestCol = worksheet.get_highest_column()
                # get next column along after last column - in which to write country code
                countryCol = highestCol
                worksheet.cell(row=0, column=countryCol).value = 'Country'
                # add the country code beneath the new heading
                print 'Processing file for {0}'.format(country)
                for rowNum in range(1, highestRow):
                    worksheet.cell(row=rowNum, column=countryCol).value = country
                workbook.save(os.path.join(filePath, outputFile.format(country, studyMonth)))
                print 'Worksheet for {0} in {1} updated and saved.'.format(country, studyMonth)
            else:
                print 'Error - File for {0} in {1} not found.'.format(country, studyMonth)
        except Exception, e:
            print 'Error: file for {0} not processed. {1}.'.format(country, e)

    print "Country code appending complete"

# #########################################################################


outputCountryCodes = addCountryCodeColumn(studyMonth, countryCodes)